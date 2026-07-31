"""
Generate 5 additional Excel tables that extend the TechNova schema:
  11. Customer_Contracts        — 24 cols, ~120 rows, FK: customers, employees
  12. Sales_Pipeline            — 24 cols, ~85 rows,  FK: customers, employees
  13. Product_Feature_Adoption  — 24 cols, ~450 rows, FK: customers, products_services, employees
  14. Audit_Findings            — 25 cols, ~75 rows,  FK: departments, employees
  15. OnCall_Rotations          — 24 cols, ~200 rows, FK: employees, products_services, departments

Output format mirrors the existing files: each .xlsx has two sheets —
`Schema_Notes` (header + brief column documentation) and `<TableName>` (data).
The duckdb_loader's _pick_data_sheet() will then auto-pick the data sheet.

Deterministic via random.seed(42). Synthetic but referentially valid against
the live data in `data/db/technova.duckdb`.
"""

from __future__ import annotations

import random
from datetime import date, timedelta
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill

random.seed(42)
DUCKDB_PATH = Path("data/db/technova.duckdb")
OUT_DIR = Path("Structured data")
OUT_DIR.mkdir(exist_ok=True)


def load_refs() -> dict:
    """Read ID lists + key columns from existing tables for referentially-valid FKs."""
    con = duckdb.connect(str(DUCKDB_PATH), read_only=True)
    refs = {
        "employees": con.execute(
            "SELECT employee_id, first_name, last_name, level, department_id, "
            "job_title, employment_status FROM employees ORDER BY employee_id"
        ).df(),
        "customers": con.execute(
            "SELECT customer_id, customer_name, tier, region, country, industry, "
            "arr_inr_lakhs, account_status, account_manager_employee_id FROM customers"
        ).df(),
        "services": con.execute(
            "SELECT service_id, service_name, domain, owner_department_id, "
            "criticality_tier, uptime_sla_percent FROM products_services"
        ).df(),
        "departments": con.execute(
            "SELECT department_id, department_name, primary_location FROM departments"
        ).df(),
        "vendors": con.execute(
            "SELECT vendor_id, vendor_name, category FROM vendors"
        ).df(),
    }
    con.close()
    return refs


def write_xlsx(filename: str, table_name: str, df: pd.DataFrame, schema_notes: list[tuple[str, str]]) -> None:
    """Write the dual-sheet xlsx (Schema_Notes + table_name sheet) matching the existing
    convention so duckdb_loader._pick_data_sheet picks the right one."""
    path = OUT_DIR / filename
    wb = openpyxl.Workbook()
    # Schema_Notes sheet
    ws = wb.active
    ws.title = "Schema_Notes"
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9E1F2")
    ws["A1"] = f"TechNova Inc.  |  {filename.split('.')[0]}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A3"] = "Column"
    ws["B3"] = "Notes"
    for c in ("A3", "B3"):
        ws[c].font = bold
        ws[c].fill = fill
    for i, (col, note) in enumerate(schema_notes, start=4):
        ws[f"A{i}"] = col
        ws[f"B{i}"] = note
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 80

    # Data sheet
    data_ws = wb.create_sheet(table_name)
    data_ws.append(list(df.columns))
    for cell in data_ws[1]:
        cell.font = bold
        cell.fill = fill
    for row in df.itertuples(index=False):
        data_ws.append(list(row))

    # Auto-width on data
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), max((len(str(v)) for v in df[col_name].astype(str)), default=10))
        data_ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"].width = min(max_len + 2, 40)

    wb.save(path)
    print(f"  wrote {path}  ({len(df)} rows, {len(df.columns)} cols)")


def fy_quarter(d: date) -> str:
    """India FY: Apr-Mar. Q1 = Apr-Jun, Q2 = Jul-Sep, Q3 = Oct-Dec, Q4 = Jan-Mar."""
    # Determine FY (Apr-Mar)
    if d.month >= 4:
        fy_start = d.year
        fy_end = d.year + 1
    else:
        fy_start = d.year - 1
        fy_end = d.year
    fy_label = f"FY{fy_start % 100:02d}-{fy_end % 100:02d}"
    if d.month in (4, 5, 6):
        return f"Q1 {fy_label}"
    if d.month in (7, 8, 9):
        return f"Q2 {fy_label}"
    if d.month in (10, 11, 12):
        return f"Q3 {fy_label}"
    return f"Q4 {fy_label}"


# ============================================================================
# 11. Customer_Contracts
# ============================================================================
def build_customer_contracts(refs: dict) -> pd.DataFrame:
    customers = refs["customers"]
    employees = refs["employees"]
    # Each Active customer has 1-3 contracts (MSA + amendments). Churned have 1.
    rows = []
    cid_counter = 80001
    for _, c in customers.iterrows():
        is_active = c["account_status"] == "Active"
        n = random.choices([1, 2, 3], weights=[5, 4, 2])[0] if is_active else 1
        msa_id = f"MSA-{c['customer_id']}-{random.randint(2020, 2024)}"
        contract_start = date(random.randint(2020, 2024), random.randint(1, 12), random.randint(1, 28))
        for k in range(n):
            ctype = "MSA" if k == 0 else random.choice(["Amendment", "SOW", "Renewal"])
            eff = contract_start + timedelta(days=k * random.randint(60, 365))
            term_months = random.choice([12, 24, 36])
            expiry = eff + timedelta(days=term_months * 30)
            base_acv = round(float(c["arr_inr_lakhs"]) * random.uniform(0.85, 1.05), 2) if k == 0 else round(float(c["arr_inr_lakhs"]) * random.uniform(0.10, 0.40), 2)
            am_id = int(c["account_manager_employee_id"]) if pd.notna(c["account_manager_employee_id"]) else int(employees["employee_id"].sample(1).iloc[0])
            rows.append({
                "contract_id": cid_counter,
                "customer_id": int(c["customer_id"]),
                "msa_reference": msa_id,
                "contract_type": ctype,
                "contract_title": f"{ctype} - {c['customer_name']}" + (f" v{k}" if k > 0 else ""),
                "effective_date": eff.isoformat(),
                "expiry_date": expiry.isoformat(),
                "term_months": term_months,
                "auto_renew": random.choice(["Yes", "Yes", "No"]),
                "notice_period_days": random.choice([30, 60, 90]),
                "payment_terms": random.choice(["Net 30", "Net 45", "Net 60", "Net 90"]),
                "currency": random.choice(["INR", "USD", "EUR"]) if c["region"] != "APAC" else "INR",
                "base_acv_inr_lakhs": base_acv,
                "discount_pct": round(random.uniform(0, 25), 2),
                "annual_uplift_pct": random.choice([3.0, 4.0, 5.0, 7.0, 0.0]),
                "governing_law": "India" if c["region"] == "APAC" else random.choice(["Delaware", "England & Wales", "Singapore"]),
                "jurisdiction": random.choice(["Bangalore", "Mumbai", "Delhi", "Singapore", "London", "New York"]),
                "data_processing_addendum_signed": random.choice(["Yes", "Yes", "Yes", "No"]),
                "dpa_signed_date": (eff + timedelta(days=random.randint(0, 30))).isoformat() if random.random() > 0.15 else None,
                "liability_cap_inr_lakhs": round(base_acv * random.choice([1.0, 1.5, 2.0, 3.0]), 2),
                "termination_for_convenience": random.choice(["Yes", "No", "No"]),
                "service_credit_floor_pct": random.choice([5.0, 10.0, 15.0, 20.0]),
                "signed_by_technova_employee_id": am_id,
                "status": "Active" if expiry >= date(2026, 4, 22) else random.choice(["Expired", "In Renewal"]),
                "last_amendment_date": eff.isoformat() if k == 0 else (eff + timedelta(days=random.randint(0, 90))).isoformat(),
            })
            cid_counter += 1
    return pd.DataFrame(rows)


# ============================================================================
# 12. Sales_Pipeline
# ============================================================================
def build_sales_pipeline(refs: dict) -> pd.DataFrame:
    employees = refs["employees"]
    customers = refs["customers"]
    sales_emps = employees[employees["job_title"].fillna("").str.contains("Sales|Account|Revenue", case=False, regex=True)]
    if len(sales_emps) < 5:
        sales_emps = employees[employees["level"].isin(["L3", "L4", "L5", "L6"])]
    sc_emps = employees[employees["job_title"].fillna("").str.contains("Solution|Architect|Consultant", case=False, regex=True)]
    if len(sc_emps) < 3:
        sc_emps = employees[employees["level"].isin(["L4", "L5", "L6"])]
    rows = []
    oid = 600001
    # Open + closed mix
    open_count, won_count, lost_count = 50, 25, 10
    stages_open = ["Discovery", "Qualified", "Proposal", "Negotiation"]
    industries = ["BFSI", "Healthcare", "Retail", "Manufacturing", "Technology", "Media", "Energy", "Telecom"]
    countries = ["India", "USA", "UK", "Singapore", "Germany", "Japan", "UAE", "Australia"]
    competitors = ["Salesforce", "ServiceNow", "Workday", "Snowflake", "Databricks", "Internal Build", "Status Quo", None]
    today = date(2026, 4, 22)

    def make(stage: str, lost: bool = False, won: bool = False):
        nonlocal oid
        is_existing = random.random() < 0.45
        c = customers.sample(1).iloc[0] if is_existing else None
        ae = sales_emps.sample(1).iloc[0]
        sc = sc_emps.sample(1).iloc[0]
        expected = round(random.uniform(20, 1500), 2)
        prob = {
            "Discovery": 10, "Qualified": 25, "Proposal": 50, "Negotiation": 75,
            "Closed-Won": 100, "Closed-Lost": 0,
        }[stage]
        days_offset = random.randint(-180, 180) if not (lost or won) else random.randint(-365, -1)
        exp_close = today + timedelta(days=days_offset)
        created = exp_close - timedelta(days=random.randint(30, 240))
        last_act = created + timedelta(days=random.randint(1, max(1, (today - created).days)))
        deal_type = "Renewal" if (is_existing and random.random() < 0.3) else random.choice(["New Logo", "New Logo", "Expansion"])
        rows.append({
            "opportunity_id": oid,
            "opportunity_name": f"{c['customer_name'] if c is not None else random.choice(['Acme Corp', 'Beacon Systems', 'Cobalt Industries', 'Delta Logistics', 'Echo Telecom', 'Fjord Capital', 'Gravity Media', 'Helios Healthcare'])} - {deal_type}",
            "customer_id": int(c["customer_id"]) if c is not None else None,
            "prospect_name": None if c is not None else f"Prospect-{oid}",
            "account_executive_employee_id": int(ae["employee_id"]),
            "solution_consultant_employee_id": int(sc["employee_id"]),
            "stage": stage,
            "expected_acv_inr_lakhs": expected,
            "weighted_acv_inr_lakhs": round(expected * prob / 100, 2),
            "probability_pct": prob,
            "forecast_category": random.choice(["Commit", "Best Case", "Pipeline", "Omitted"]) if not (lost or won) else ("Closed Won" if won else "Closed Lost"),
            "expected_close_date": exp_close.isoformat(),
            "actual_close_date": exp_close.isoformat() if (won or lost) else None,
            "created_date": created.isoformat(),
            "last_activity_date": last_act.isoformat(),
            "lead_source": random.choice(["Inbound Web", "Outbound SDR", "Partner Referral", "Field Event", "Customer Referral"]),
            "tier_intent": random.choice(["Tier 1", "Tier 2", "Tier 3"]),
            "region": c["region"] if c is not None else random.choice(["APAC", "North America", "Europe", "MEA"]),
            "country": c["country"] if c is not None else random.choice(countries),
            "industry": c["industry"] if c is not None else random.choice(industries),
            "deal_type": deal_type,
            "competitor_in_play": random.choice(competitors),
            "quarterly_period": fy_quarter(exp_close),
            "lost_reason": random.choice(["Price", "Product Gap", "Competitor", "Timing", "No Decision"]) if lost else None,
            "lost_to_competitor": random.choice(["Salesforce", "ServiceNow", "Internal Build"]) if lost and random.random() < 0.6 else None,
            "notes": random.choice(["Strong product-fit, awaiting CFO sign-off.", "Proof-of-value live.", "Procurement freeze.", "Awaiting security review.", "Champion left customer."]),
        })
        oid += 1

    for _ in range(open_count):
        make(random.choices(stages_open, weights=[3, 4, 5, 4])[0])
    for _ in range(won_count):
        make("Closed-Won", won=True)
    for _ in range(lost_count):
        make("Closed-Lost", lost=True)
    return pd.DataFrame(rows)


# ============================================================================
# 13. Product_Feature_Adoption
# ============================================================================
def build_feature_adoption(refs: dict) -> pd.DataFrame:
    customers = refs["customers"]
    services = refs["services"]
    employees = refs["employees"]
    rows = []
    adopt_id = 700001
    feature_catalog = [
        ("Real-time Dashboards", "Core", "2023-04-15"),
        ("Single Sign-On (SSO)", "Core", "2023-06-01"),
        ("Audit Logs Export", "Core", "2023-08-22"),
        ("AI-Powered Anomaly Detection", "AI/ML", "2024-09-10"),
        ("AI Co-pilot (Beta)", "AI/ML", "2025-02-01"),
        ("Custom Workflows", "Advanced", "2024-01-12"),
        ("API Webhooks", "Integration", "2023-09-20"),
        ("Salesforce Connector", "Integration", "2024-03-15"),
        ("Workday Connector", "Integration", "2024-07-01"),
        ("Multi-region Replication", "Advanced", "2025-01-15"),
        ("Customer-managed Keys (BYOK)", "Advanced", "2025-04-01"),
        ("Predictive Forecasting (ML)", "AI/ML", "2024-11-08"),
    ]
    for _, c in customers.iterrows():
        if c["account_status"] not in ("Active", "Expansion"):
            n_features = random.randint(1, 3)
        else:
            n_features = random.randint(3, 8)
        chosen = random.sample(feature_catalog, n_features)
        for fname, fcat, frelease in chosen:
            services_in_domain = services[services["domain"].isin(["Analytics Engine", "AI Services", "User Management", "Data Processing", "Integration Hub"])]
            svc = services_in_domain.sample(1).iloc[0]
            mau = random.randint(0, int(c["arr_inr_lakhs"] / 5)) if random.random() > 0.05 else 0
            wau = int(mau * random.uniform(0.4, 0.8))
            dau = int(wau * random.uniform(0.3, 0.7))
            adoption = random.choices(["Active", "Active", "Active", "Trialing", "Disabled", "Deprecated"], weights=[5, 5, 5, 3, 2, 1])[0]
            activation = (date.fromisoformat(frelease) + timedelta(days=random.randint(15, 400))).isoformat()
            health = random.randint(20, 100) if adoption == "Active" else random.randint(0, 60)
            churn_risk = "Yes" if (health < 50 and adoption == "Active") else "No"
            expansion = "Yes" if (health > 80 and fcat in ("AI/ML", "Advanced")) else "No"
            am_id = int(c["account_manager_employee_id"]) if pd.notna(c["account_manager_employee_id"]) else int(employees.sample(1).iloc[0]["employee_id"])
            rows.append({
                "adoption_id": adopt_id,
                "customer_id": int(c["customer_id"]),
                "service_id": int(svc["service_id"]),
                "feature_name": fname,
                "feature_category": fcat,
                "feature_release_date": frelease,
                "adoption_status": adoption,
                "activation_date": activation if adoption != "Disabled" else None,
                "last_used_date": (date(2026, 4, 22) - timedelta(days=random.randint(0, 90))).isoformat() if adoption == "Active" else (date(2026, 4, 22) - timedelta(days=random.randint(120, 400))).isoformat(),
                "usage_frequency": random.choice(["Daily", "Daily", "Weekly", "Monthly", "Rarely"]) if adoption == "Active" else "Rarely",
                "mau_last_30d": mau,
                "wau_last_7d": wau,
                "dau": dau,
                "configuration_completeness_pct": round(random.uniform(40, 100), 1),
                "integrations_enabled_count": random.randint(0, 5),
                "api_calls_last_30d": random.randint(0, 1_000_000) if mau > 0 else 0,
                "error_rate_pct": round(random.uniform(0, 3.5), 2),
                "customer_satisfaction_score": round(random.uniform(3.0, 9.5), 1),
                "nps_rating_date": (date(2026, 4, 22) - timedelta(days=random.randint(7, 180))).isoformat(),
                "enabled_by_employee_id": am_id,
                "health_score": health,
                "churn_risk_flag": churn_risk,
                "expansion_opportunity_flag": expansion,
                "feedback_count_last_quarter": random.randint(0, 8),
            })
            adopt_id += 1
    return pd.DataFrame(rows)


# ============================================================================
# 14. Audit_Findings
# ============================================================================
def build_audit_findings(refs: dict) -> pd.DataFrame:
    departments = refs["departments"]
    employees = refs["employees"]
    rows = []
    fid = 80001
    audit_engagements = [
        ("SOC 2 Type II", "Deloitte Touche Tohmatsu", "2025-Q3", date(2025, 4, 1), date(2025, 9, 30), date(2025, 12, 15)),
        ("ISO 27001", "BSI India", "2025-Q4", date(2025, 7, 1), date(2025, 12, 31), date(2026, 2, 28)),
        ("Internal Audit FY2025-26 Q1", "Internal", "2025-Q1", date(2025, 4, 1), date(2025, 6, 30), date(2025, 8, 15)),
        ("Internal Audit FY2025-26 Q3", "Internal", "2025-Q3", date(2025, 10, 1), date(2025, 12, 31), date(2026, 2, 1)),
        ("Penetration Test (External)", "NCC Group", "2025-Q4", date(2025, 11, 1), date(2025, 11, 30), date(2025, 12, 20)),
        ("DPDP Compliance Audit", "Khaitan & Co Legal", "2026-Q1", date(2026, 1, 1), date(2026, 3, 31), date(2026, 4, 10)),
        ("Statutory Audit FY2025-26", "Deloitte Touche Tohmatsu", "2026-Q1", date(2025, 4, 1), date(2026, 3, 31), date(2026, 4, 15)),
    ]
    finding_categories = [
        ("Control Gap", "Medium"), ("Documentation", "Low"), ("Process", "Medium"),
        ("Technical Debt", "Medium"), ("Vulnerability", "High"), ("Vulnerability", "Critical"),
        ("Configuration", "Low"), ("Access Management", "High"), ("Data Handling", "High"),
        ("Vendor Management", "Medium"),
    ]
    statuses_w = [("Open", 2), ("In Progress", 4), ("Remediated", 5), ("Accepted Risk", 1), ("Closed", 3)]
    statuses = [s for s, w in statuses_w for _ in range(w)]

    for engagement_name, firm, eng_q, start, end, report in audit_engagements:
        n_findings = random.randint(6, 14)
        for _ in range(n_findings):
            category, sev = random.choice(finding_categories)
            dept = departments.sample(1).iloc[0]
            owner = employees[(employees["department_id"] == dept["department_id"]) & (employees["level"].isin(["L4", "L5", "L6", "L7"]))]
            owner = owner.sample(1).iloc[0] if not owner.empty else employees.sample(1).iloc[0]
            status = random.choice(statuses) if sev != "Critical" else random.choice(["Open", "In Progress", "Remediated"])
            due = report + timedelta(days=random.randint(30, 180))
            completion = (due - timedelta(days=random.randint(0, 60))).isoformat() if status in ("Remediated", "Closed") else None
            is_repeat = random.random() < 0.18
            ipo_blocker = "Yes" if (sev in ("Critical", "High") and status in ("Open", "In Progress") and "SOC 2" in engagement_name + " ISO 27001 Penetration") else "No"
            ipo_blocker = "Yes" if (sev == "Critical" and status != "Closed") else ipo_blocker
            cust_visibility = random.choice(["Internal", "Internal", "Internal", "Customer-Audit-Visible", "Investor-Visible"])
            if "SOC 2" in engagement_name or "ISO 27001" in engagement_name:
                cust_visibility = random.choice(["Customer-Audit-Visible", "Investor-Visible"])
            rows.append({
                "finding_id": fid,
                "audit_engagement": engagement_name,
                "audit_type": engagement_name.split(" ")[0] + (" " + engagement_name.split(" ")[1] if "Type" in engagement_name else ""),
                "audit_year": end.year,
                "audit_quarter": eng_q,
                "auditor_firm": firm,
                "auditor_lead_name": random.choice(["Priya Iyer", "Karthik Menon", "Sarah Hopkins", "Ahmed Khan", "Rachel Cohen", "Naveen Pillai"]),
                "audit_period_start": start.isoformat(),
                "audit_period_end": end.isoformat(),
                "report_date": report.isoformat(),
                "department_id": int(dept["department_id"]),
                "finding_category": category,
                "severity": sev,
                "finding_summary": f"{category} identified in {dept['department_name']} ({sev} severity)",
                "control_id": f"{random.choice(['CC', 'A', 'P'])}{random.choice(['6.1', '7.2', '5.3', '8.1', '4.4'])}",
                "status": status,
                "assigned_owner_employee_id": int(owner["employee_id"]),
                "due_date": due.isoformat(),
                "completion_date": completion,
                "remediation_cost_inr_lakhs": round(random.uniform(0.5, 25.0), 2),
                "is_repeat_finding": "Yes" if is_repeat else "No",
                "previous_finding_id": fid - random.randint(2, 30) if is_repeat and fid > 80035 else None,
                "customer_visibility": cust_visibility,
                "ipo_blocker_flag": ipo_blocker,
                "evidence_artifacts": random.choice(["JIRA-AUD-#####", "Confluence page", "Email trail", "Test report PDF", "Screen recording"]),
            })
            fid += 1
    return pd.DataFrame(rows)


# ============================================================================
# 15. OnCall_Rotations
# ============================================================================
def build_oncall_rotations(refs: dict) -> pd.DataFrame:
    services = refs["services"]
    employees = refs["employees"]
    departments = refs["departments"]
    eng_employees = employees[employees["level"].isin(["L3", "L4", "L5", "L6"])]
    rows = []
    rid = 90001
    # Last 24 weeks of rotations for ~10 critical services
    critical_services = services[services["criticality_tier"].isin(["Critical", "High"])].head(15)
    today = date(2026, 4, 22)
    for week_offset in range(0, 24):
        week_start = today - timedelta(days=week_offset * 7 + (today.weekday()))
        week_end = week_start + timedelta(days=6)
        for _, svc in critical_services.iterrows():
            for role in ["Primary", "Secondary"]:
                emp_pool = eng_employees[eng_employees["department_id"] == svc["owner_department_id"]]
                if emp_pool.empty:
                    emp_pool = eng_employees
                emp = emp_pool.sample(1).iloc[0]
                pages = random.randint(0, 14) if role == "Primary" else random.randint(0, 4)
                ack = max(0, pages - random.randint(0, 2))
                avg_resp = round(random.uniform(2, 18), 1) if role == "Primary" else round(random.uniform(5, 25), 1)
                mttr = round(avg_resp + random.uniform(15, 240), 1) if pages > 0 else 0.0
                stipend = 5000 if role == "Primary" else 2500
                holiday = random.choice([0, 0, 0, 1500, 3000])
                swap_count = random.choice([0, 0, 0, 1, 2])
                covered_by = int(eng_employees.sample(1).iloc[0]["employee_id"]) if swap_count > 0 else None
                burnout = min(100, int(pages * 5 + (week_offset == 0) * 10 + random.randint(0, 30)))
                rows.append({
                    "rotation_id": rid,
                    "service_id": int(svc["service_id"]),
                    "service_name": svc["service_name"],
                    "employee_id": int(emp["employee_id"]),
                    "employee_name": f"{emp['first_name']} {emp['last_name']}",
                    "role": role,
                    "department_id": int(svc["owner_department_id"]),
                    "team_name": departments[departments["department_id"] == svc["owner_department_id"]].iloc[0]["department_name"],
                    "rotation_start_date": week_start.isoformat(),
                    "rotation_end_date": week_end.isoformat(),
                    "shift_type": random.choice(["Business Hours", "After Hours", "Weekend", "24/7"]),
                    "timezone": random.choice(["IST", "PST", "UTC", "GMT"]),
                    "pages_received_count": pages,
                    "acknowledged_within_sla_count": ack,
                    "avg_response_time_minutes": avg_resp,
                    "mean_time_to_resolve_minutes": mttr,
                    "weekly_stipend_inr": stipend,
                    "holiday_premium_paid_inr": holiday,
                    "swap_count": swap_count,
                    "covered_by_employee_id": covered_by,
                    "burnout_risk_score": burnout,
                    "status": "Completed" if week_offset > 0 else "Active",
                    "handoff_notes": random.choice(["Clean week, no escalations.", "Two SEV-3 mitigations.", "DB latency spike resolved.", "Vendor dependency degraded.", "Alarm tuning needed."]),
                    "created_by_employee_id": int(emp_pool.sample(1).iloc[0]["employee_id"]),
                })
                rid += 1
    return pd.DataFrame(rows)


# ============================================================================
# Main
# ============================================================================
def main():
    print("Loading reference data from DuckDB...")
    refs = load_refs()
    print(f"  employees={len(refs['employees'])}  customers={len(refs['customers'])}  "
          f"services={len(refs['services'])}  departments={len(refs['departments'])}  "
          f"vendors={len(refs['vendors'])}")
    print()

    builders = [
        ("11_Customer_Contracts.xlsx", "Customer_Contracts", build_customer_contracts, [
            ("contract_id", "PK. Contract identifier."),
            ("customer_id", "FK -> customers.customer_id"),
            ("msa_reference", "Master Services Agreement reference (e.g. 'MSA-7045-2023')"),
            ("contract_type", "MSA / Amendment / SOW / Renewal"),
            ("base_acv_inr_lakhs", "Annual Contract Value at base pricing"),
            ("liability_cap_inr_lakhs", "Cap on TechNova's liability under this contract"),
            ("data_processing_addendum_signed", "Whether DPA is in place (Yes/No)"),
            ("status", "Active / Expired / In Renewal / Terminated"),
        ]),
        ("12_Sales_Pipeline.xlsx", "Sales_Pipeline", build_sales_pipeline, [
            ("opportunity_id", "PK. Opportunity identifier."),
            ("customer_id", "FK -> customers.customer_id (NULL for new prospects)"),
            ("account_executive_employee_id", "FK -> employees.employee_id (the AE owning the deal)"),
            ("solution_consultant_employee_id", "FK -> employees.employee_id"),
            ("stage", "Discovery / Qualified / Proposal / Negotiation / Closed-Won / Closed-Lost"),
            ("expected_acv_inr_lakhs", "Forecast ACV if won"),
            ("weighted_acv_inr_lakhs", "expected_acv * probability/100 — for forecasting"),
            ("forecast_category", "Commit / Best Case / Pipeline / Omitted / Closed Won / Closed Lost"),
        ]),
        ("13_Product_Feature_Adoption.xlsx", "Product_Feature_Adoption", build_feature_adoption, [
            ("adoption_id", "PK. Per-customer-per-feature adoption record."),
            ("customer_id", "FK -> customers.customer_id"),
            ("service_id", "FK -> products_services.service_id"),
            ("feature_name", "Named feature (e.g. 'AI Co-pilot (Beta)')"),
            ("feature_category", "Core / Advanced / AI/ML / Integration"),
            ("mau_last_30d", "Monthly active users from this customer"),
            ("health_score", "Derived 0-100 health score"),
            ("churn_risk_flag", "Yes if low health AND active customer"),
            ("expansion_opportunity_flag", "Yes if high health on AI/ML or Advanced features"),
        ]),
        ("14_Audit_Findings.xlsx", "Audit_Findings", build_audit_findings, [
            ("finding_id", "PK. Finding identifier."),
            ("audit_engagement", "Named engagement (SOC 2 Type II, ISO 27001, Internal Audit, etc.)"),
            ("auditor_firm", "External firm or 'Internal'"),
            ("severity", "Critical / High / Medium / Low / Observation"),
            ("status", "Open / In Progress / Remediated / Accepted Risk / Closed"),
            ("department_id", "FK -> departments.department_id"),
            ("assigned_owner_employee_id", "FK -> employees.employee_id"),
            ("ipo_blocker_flag", "Yes if this finding blocks IPO readiness (Critical/High + open)"),
            ("is_repeat_finding", "Yes if same root cause as a previous finding"),
        ]),
        ("15_OnCall_Rotations.xlsx", "OnCall_Rotations", build_oncall_rotations, [
            ("rotation_id", "PK. Per-employee-per-week rotation."),
            ("service_id", "FK -> products_services.service_id"),
            ("employee_id", "FK -> employees.employee_id"),
            ("role", "Primary / Secondary / Manager Escalation"),
            ("weekly_stipend_inr", "5000 primary, 2500 secondary per OnCall_Runbook policy"),
            ("avg_response_time_minutes", "Average ack time during this rotation"),
            ("mean_time_to_resolve_minutes", "MTTR for incidents during this rotation"),
            ("burnout_risk_score", "0-100 derived from page volume + tenure"),
            ("covered_by_employee_id", "FK -> employees.employee_id (NULL unless rotation was swapped)"),
        ]),
    ]

    for filename, table_name, builder, schema in builders:
        print(f"Building {filename}...")
        df = builder(refs)
        write_xlsx(filename, table_name, df, schema)


if __name__ == "__main__":
    main()
